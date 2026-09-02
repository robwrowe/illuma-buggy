"""Load labeled effect rows from Op_Codes_Captured.xlsx into TrialRow records.

Packet model (see docs/ble-packets-details/fields/length-byte.md): the byte after
the Disney identifier E9 is a length byte (`payload_length = length_byte + 2`),
not an opcode/category selector. `op_code` is carried as a display label only so
a reviewer can cross-reference `docs/ble-packets-details/op-codes/` raw samples.

Field-role names match web/src/lib/ble/tailBuilder.ts:
  env / fixed / len / timing / format / color / tail / vib
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

COMPANY_ID = bytes((0x83, 0x01))
ENV_BYTES = {0xE1, 0xE2}
DISNEY_ID = 0xE9
VIB_HIGH_NIBBLE = 0xB
TAIL_COL_COUNT = 17  # T00 .. T16

# Closed effect vocabulary from the xlsx INSTRUCTIONS sheet, plus labels from
# the second labeled sheet (Solid / SW Twinkle / Flicker Chase).
XLSX_EFFECT_LABELS = (
    "Chase",
    "Shimmer",
    "Flicker",
    "Pulse",
    "Cycle",
    "Strobe",
    "Heartbeat",
    "Cross-saw",
    "Cross-fade",
    "Unique",
    "Circle",
    "Glow",
    "Solid",
    "SW Twinkle",
    "Flicker Chase",
)

SOURCE_OP_CODES_CAPTURED = "op_codes_captured"
SOURCE_SECOND_LABELED = "second_labeled_sheet"
SOURCE_KEYED_NOTES = "groundtruth_keyed_notes"
SOURCE_BUILDER = "builder"


def normalize_hex(value: str) -> str:
    return re.sub(r"[^0-9a-fA-F]", "", str(value or "")).upper()


def hex_to_bytes(value: str) -> bytes:
    h = normalize_hex(value)
    if not h:
        return b""
    if len(h) % 2:
        h = h[:-1]
    return bytes.fromhex(h)


def ensure_hex_full(hex_str: str) -> tuple[str, bool]:
    """Return (8301-prefixed hex, envelope_assumed).

    Payload-only E9-leading rows get 8301E100 prepended so POST /show can
    take them. /send is not used for capture: it blocks for the hold time,
    so the webcam would start after the effect had already ended.
    """
    h = normalize_hex(hex_str)
    if not h:
        return "", False
    if h.startswith("8301"):
        return h, False
    if h.startswith("E1") or h.startswith("E2"):
        return "8301" + h, False
    if h.startswith("E9"):
        return "8301E100" + h, True
    return "8301E100" + h, True


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9#]+", "", str(value or "").strip().lower())


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text if text else None


def _cell_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cell_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        if text.lower().startswith("0x"):
            return int(text, 16)
        return int(text, 10)
    except ValueError:
        try:
            return int(text, 16)
        except ValueError:
            return None


def fs_safe(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "row"


@dataclass
class ZoneLayoutHint:
    five_zones: str | None = None
    sync: str | None = None
    layout: str | None = None
    direction: str | None = None
    n_zones: int | None = None
    cycle_length: float | None = None
    n_cycles: float | None = None


@dataclass
class DecodedPacket:
    """Structural decode of hex_full. Does not bit-slice tail/animation bytes."""

    raw: bytes
    after_company_id: bytes
    e9_index: int | None
    length_byte: int | None
    derived_payload_length: int | None
    actual_payload_length: int | None
    length_mismatch: bool
    vibration_byte: int | None
    vibration_nibble: int | None
    parts: list[dict[str, Any]] = field(default_factory=list)


def decode_hex_structure(hex_full: str) -> DecodedPacket:
    """Strip 8301, find E9, read the length byte, detect a trailing vib byte.

    Layout after the company ID, matching tailBuilder.py / tailBuilder.ts:
      env (0xE1/0xE2) → envPad (0x00) → e9 (0xE9) → sub/len → pad → tb →
      fmt → color block → tail → optional vib (high nibble 0xB).
    """
    raw = hex_to_bytes(hex_full)
    after = raw[2:] if raw[:2] == COMPANY_ID else raw
    e9_index = after.find(bytes((DISNEY_ID,)))
    if e9_index < 0:
        return DecodedPacket(
            raw=raw,
            after_company_id=after,
            e9_index=None,
            length_byte=None,
            derived_payload_length=None,
            actual_payload_length=None,
            length_mismatch=False,
            vibration_byte=None,
            vibration_nibble=None,
        )

    disney = after[e9_index:]
    length_byte = disney[1] if len(disney) >= 2 else None
    derived = (length_byte + 2) if length_byte is not None else None
    actual = len(disney)
    mismatch = bool(derived is not None and derived != actual)

    vib_byte = None
    vib_nibble = None
    if disney and (disney[-1] >> 4) == VIB_HIGH_NIBBLE:
        vib_byte = disney[-1]
        vib_nibble = disney[-1] & 0x0F

    parts: list[dict[str, Any]] = []
    i = 0
    if after and after[0] in ENV_BYTES:
        parts.append({"id": "env", "role": "env", "byte": after[0]})
        i = 1
        if i < len(after) and after[i] == 0x00:
            parts.append({"id": "envPad", "role": "fixed", "byte": 0x00})
            i += 1
    if i < len(after) and after[i] == DISNEY_ID:
        parts.append({"id": "e9", "role": "fixed", "byte": DISNEY_ID})
        i += 1
    if i < len(after) and length_byte is not None:
        parts.append({"id": "sub", "role": "len", "byte": after[i]})
        i += 1
    if i < len(after) and after[i] == 0x00:
        parts.append({"id": "pad", "role": "fixed", "byte": 0x00})
        i += 1
    if i < len(after):
        parts.append({"id": "tb", "role": "timing", "byte": after[i]})
        i += 1
    if i < len(after):
        parts.append({"id": "fmt", "role": "format", "byte": after[i]})
        i += 1
    # Remaining bytes: color + tail + optional vib. We do not split color vs tail
    # from hex (xlsx T00..T16 / F1 / # carry that); just mark vib if present.
    remaining = list(after[i:])
    if vib_byte is not None and remaining and remaining[-1] == vib_byte:
        remaining = remaining[:-1]
        parts.append({"id": "payload_mid", "role": "color", "bytes": remaining})
        parts.append({"id": "vib", "role": "vib", "byte": vib_byte})
    else:
        parts.append({"id": "payload_mid", "role": "tail", "bytes": remaining})

    return DecodedPacket(
        raw=raw,
        after_company_id=after,
        e9_index=e9_index,
        length_byte=length_byte,
        derived_payload_length=derived,
        actual_payload_length=actual,
        length_mismatch=mismatch,
        vibration_byte=vib_byte,
        vibration_nibble=vib_nibble,
        parts=parts,
    )


@dataclass
class TrialRow:
    sheet: str
    row_id: str
    row_index: int
    op_code: str | None
    length_byte: int | None
    derived_payload_length: int | None
    color_count: int | None
    color_format_1: str | None
    color_format_2: str | None
    effect_label: str | None
    description: str | None
    hex_full: str
    hex_key: str
    location: str | None
    show: str | None
    date: str | None
    tail_start_index: int | None
    tail_bytes: list[tuple[int, str]]
    vibration_byte: str | None
    notes: list[str] = field(default_factory=list)
    capture_source_row_id: str | None = None
    duplicate_count: int = 1
    decoded: DecodedPacket | None = None
    zone_layout_hint: ZoneLayoutHint = field(default_factory=ZoneLayoutHint)
    source_sheet_kind: str = SOURCE_OP_CODES_CAPTURED
    envelope_assumed: bool = False
    zone_layout_downgraded: bool = False
    expected_colors: list = field(default_factory=list)

    @property
    def sheet_safe(self) -> str:
        return fs_safe(self.sheet)

    @property
    def row_id_safe(self) -> str:
        return fs_safe(self.row_id)

    def tail_bytes_summary(self) -> str:
        if not self.tail_bytes:
            return ""
        return " ".join(f"T{idx:02d}={val}" for idx, val in self.tail_bytes)


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "opcode": ("opcode", "op_code", "op"),
    "length": ("length", "len", "lengthbyte"),
    "color_count": ("#", "n", "colors", "colorcount", "numcolors"),
    "f1": ("f1", "colorformat1", "fmt1"),
    "f2": ("f2", "colorformat2", "fmt2"),
    "effect": ("effect", "effectlabel", "label"),
    "description": ("description", "desc", "notes"),
    "hex": ("hex", "hexfull", "payload", "advertisement", "effectivecode"),
    "location": ("location", "loc", "park"),
    "show": ("show", "showname"),
    "date": ("date", "captured", "capturedat"),
    "start": ("start", "tailstart", "tailstartindex"),
    "vib": ("vib", "vibration", "vibrationbyte"),
    "five_zones": ("5zones", "fivezones", "5zone"),
    "sync": ("sync",),
    "layout": ("layout",),
    "direction": ("direction", "dir"),
    "n_zones": ("ofzones", "numzones", "nzones", "zones", "#ofzones"),
    "cycle_length": ("cyclelength", "cyclelen"),
    "n_cycles": ("ofcycles", "ncycles", "cycles", "#ofcycles"),
}


def _build_colmap(headers: list[str]) -> dict[str, int]:
    norms = [_norm_header(h) for h in headers]
    colmap: dict[str, int] = {}
    for logical, aliases in HEADER_ALIASES.items():
        for i, n in enumerate(norms):
            if n in aliases:
                colmap[logical] = i
                break
    for i, n in enumerate(norms):
        m = re.fullmatch(r"t(\d{1,2})", n)
        if m:
            colmap[f"t{int(m.group(1)):02d}"] = i
    return colmap


def _find_header_row(rows: Iterable[tuple[int, tuple[Any, ...]]]) -> tuple[int, list[str]] | None:
    for row_index, values in rows:
        headers = [_cell_str(v) or "" for v in values]
        norms = [_norm_header(h) for h in headers]
        if "hex" in norms and ("effect" in norms or "opcode" in norms or "op_code" in norms):
            return row_index, headers
        if "hex" in norms and "#" in norms:
            return row_index, headers
        if "effectivecode" in norms and "effect" in norms:
            return row_index, headers
    return None


def _sheet_rows(ws) -> list[tuple[int, tuple[Any, ...]]]:
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        out.append((i, row))
    return out


def _get(values: tuple[Any, ...], colmap: dict[str, int], key: str) -> Any:
    idx = colmap.get(key)
    if idx is None or idx >= len(values):
        return None
    return values[idx]


def _parse_trial(sheet: str, row_index: int, values: tuple[Any, ...], colmap: dict[str, int]) -> TrialRow | None:
    hex_raw = _cell_str(_get(values, colmap, "hex"))
    if not hex_raw:
        return None
    hex_key = normalize_hex(hex_raw)
    if len(hex_key) < 4:
        return None

    effect = _cell_str(_get(values, colmap, "effect"))
    length_from_xlsx = _cell_int(_get(values, colmap, "length"))
    decoded = decode_hex_structure(hex_raw)

    notes: list[str] = []
    length_byte = length_from_xlsx if length_from_xlsx is not None else decoded.length_byte
    # Length column is sometimes the hex length-byte (e.g. 0x11 / 17) and
    # sometimes already the derived payload length. Prefer the decoded E9+1 byte.
    if decoded.length_byte is not None:
        if length_from_xlsx is not None and length_from_xlsx != decoded.length_byte:
            if length_from_xlsx == decoded.derived_payload_length:
                notes.append(
                    f"xlsx Length={length_from_xlsx} looks like derived payload length; "
                    f"using decoded length_byte=0x{decoded.length_byte:02X}"
                )
            else:
                notes.append(
                    f"xlsx Length={length_from_xlsx} disagrees with decoded length_byte="
                    f"0x{decoded.length_byte:02X} (payload_length={decoded.derived_payload_length}, "
                    f"actual={decoded.actual_payload_length})"
                )
        length_byte = decoded.length_byte

    derived = (length_byte + 2) if length_byte is not None else decoded.derived_payload_length
    if decoded.length_mismatch:
        notes.append(
            f"length mismatch: length_byte+2={decoded.derived_payload_length} "
            f"but Disney payload (from E9) is {decoded.actual_payload_length} bytes"
        )

    vib_xlsx = _cell_str(_get(values, colmap, "vib"))
    if vib_xlsx is None and decoded.vibration_byte is not None:
        notes.append(
            "vibration_byte_mismatch: Vib column blank but last payload byte "
            f"high nibble is 0xB (0x{decoded.vibration_byte:02X})"
        )
    elif vib_xlsx is not None and decoded.vibration_byte is None:
        notes.append(
            f"vibration_byte_mismatch: Vib column={vib_xlsx} but last payload byte "
            "is not high-nibble 0xB"
        )

    tail_bytes: list[tuple[int, str]] = []
    for t_i in range(TAIL_COL_COUNT):
        raw = _cell_str(_get(values, colmap, f"t{t_i:02d}"))
        if raw is None:
            continue
        tail_bytes.append((t_i, raw))

    color_count = _cell_int(_get(values, colmap, "color_count"))
    if not tail_bytes:
        tail_bytes = tail_bytes_from_decoded(decoded, color_count)

    hint = ZoneLayoutHint(
        five_zones=_cell_str(_get(values, colmap, "five_zones")),
        sync=_cell_str(_get(values, colmap, "sync")),
        layout=_cell_str(_get(values, colmap, "layout")),
        direction=_cell_str(_get(values, colmap, "direction")),
        n_zones=_cell_int(_get(values, colmap, "n_zones")),
        cycle_length=_cell_float(_get(values, colmap, "cycle_length")),
        n_cycles=_cell_float(_get(values, colmap, "n_cycles")),
    )

    full, env_assumed = ensure_hex_full(hex_raw)
    if env_assumed:
        notes.append("envelope_assumed: prepended 8301E100 (payload-only hex)")

    return TrialRow(
        sheet=sheet,
        row_id=f"{sheet}:{row_index}",
        row_index=row_index,
        op_code=_cell_str(_get(values, colmap, "opcode")),
        length_byte=length_byte,
        derived_payload_length=derived,
        color_count=color_count,
        color_format_1=_cell_str(_get(values, colmap, "f1")),
        color_format_2=_cell_str(_get(values, colmap, "f2")),
        effect_label=effect,
        description=_cell_str(_get(values, colmap, "description")),
        hex_full=hex_raw if not env_assumed else full,
        hex_key=full,
        location=_cell_str(_get(values, colmap, "location")),
        show=_cell_str(_get(values, colmap, "show")),
        date=_cell_str(_get(values, colmap, "date")),
        tail_start_index=_cell_int(_get(values, colmap, "start")),
        tail_bytes=tail_bytes,
        vibration_byte=vib_xlsx,
        notes=notes,
        decoded=decoded,
        zone_layout_hint=hint,
        envelope_assumed=env_assumed,
        expected_colors=infer_expected_colors(decoded, color_count),
    )


def _iter_sheet_trials(ws, sheet_name: str) -> list[TrialRow]:
    raw_rows = _sheet_rows(ws)
    header = _find_header_row(raw_rows[:15])
    if header is None:
        return []
    header_row_index, headers = header
    colmap = _build_colmap(headers)
    if "hex" not in colmap:
        return []
    trials: list[TrialRow] = []
    for row_index, values in raw_rows:
        if row_index <= header_row_index:
            continue
        trial = _parse_trial(sheet_name, row_index, values, colmap)
        if trial is not None:
            trials.append(trial)
    return trials


@dataclass
class TrialSet:
    trials: list[TrialRow]
    by_hex: dict[str, list[TrialRow]]

    def unique_capture_trials(self) -> list[TrialRow]:
        """First row of each unique hex_full, in first-seen order."""
        seen: set[str] = set()
        out: list[TrialRow] = []
        for t in self.trials:
            if t.hex_key in seen:
                continue
            seen.add(t.hex_key)
            out.append(t)
        return out


def load_trials(xlsx_path: str | Path) -> TrialSet:
    path = Path(xlsx_path)
    if not path.is_file():
        raise FileNotFoundError(f"xlsx not found: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        efx_names = [n for n in wb.sheetnames if n.lower().startswith("efx_")]
        trials: list[TrialRow] = []
        seen_hex: set[str] = set()
        for name in efx_names:
            rows = _iter_sheet_trials(wb[name], name)
            trials.extend(rows)
            seen_hex.update(t.hex_key for t in rows)

        unique_name = next(
            (n for n in wb.sheetnames if n.lower().replace(" ", "") in {
                "absoluteuniquecode",
                "absoluteunique",
                "uniquecode",
            }),
            None,
        )
        if unique_name is None:
            unique_name = next((n for n in wb.sheetnames if "unique" in n.lower() and not n.lower().startswith("efx_")), None)
        if unique_name is not None:
            for row in _iter_sheet_trials(wb[unique_name], unique_name):
                if row.hex_key not in seen_hex:
                    trials.append(row)
                    seen_hex.add(row.hex_key)

        consumed = set(efx_names)
        if unique_name:
            consumed.add(unique_name)
        for name in wb.sheetnames:
            if name in consumed:
                continue
            extra = _iter_sheet_trials(wb[name], name)
            if not extra:
                continue
            # Header-detected second labeled sheet (Effective Code / Effect / Description).
            if extra[0].envelope_assumed or extra[0].source_sheet_kind == SOURCE_SECOND_LABELED:
                for row in extra:
                    row.source_sheet_kind = SOURCE_SECOND_LABELED
                    if not row.tail_bytes and row.decoded:
                        row.tail_bytes = tail_bytes_from_decoded(row.decoded, row.color_count)
                    trials.append(row)
    finally:
        wb.close()

    return _retag(trials)


def filter_trials(
    trial_set: TrialSet,
    *,
    sheet: str | None = None,
    limit: int | None = None,
) -> TrialSet:
    trials = trial_set.trials
    if sheet:
        want = sheet.lower()
        trials = [t for t in trials if t.sheet.lower() == want]
        if not trials:
            available = sorted({t.sheet for t in trial_set.trials})
            raise ValueError(f"no trials on sheet {sheet!r}; available: {available}")
    if limit is not None:
        unique_keys: list[str] = []
        seen: set[str] = set()
        for t in trials:
            if t.hex_key in seen:
                continue
            seen.add(t.hex_key)
            unique_keys.append(t.hex_key)
            if len(unique_keys) >= limit:
                break
        allowed = set(unique_keys)
        trials = [t for t in trials if t.hex_key in allowed]
    return _retag(trials)


def _retag(trials: list[TrialRow]) -> TrialSet:
    by_hex: dict[str, list[TrialRow]] = {}
    for t in trials:
        by_hex.setdefault(t.hex_key, []).append(t)
    for group in by_hex.values():
        source = group[0].row_id
        for t in group:
            t.capture_source_row_id = source
            t.duplicate_count = len(group)
    return TrialSet(trials=trials, by_hex=by_hex)


def merge_trial_sources(*sets: TrialSet) -> TrialSet:
    """Concatenate sources; capture once per hex_key, keep every labeled row."""
    trials: list[TrialRow] = []
    for s in sets:
        if s is None:
            continue
        trials.extend(s.trials)
    return _retag(trials)


def _payload_after_format(decoded: DecodedPacket) -> tuple[int | None, list[int]]:
    """Format byte and the bytes after it (vib stripped). Empty if undecodable."""
    after = decoded.after_company_id
    if not after:
        return None, []
    i = 0
    if after and after[0] in ENV_BYTES:
        i = 1
        if i < len(after) and after[i] == 0x00:
            i += 1
    if i < len(after) and after[i] == DISNEY_ID:
        i += 1
    if i < len(after):
        i += 1  # len
    if i < len(after) and after[i] == 0x00:
        i += 1
    if i < len(after):
        i += 1  # tb
    fmt = after[i] if i < len(after) else None
    if i < len(after):
        i += 1
    rest = list(after[i:])
    if decoded.vibration_byte is not None and rest and rest[-1] == decoded.vibration_byte:
        rest = rest[:-1]
    return fmt, rest


def infer_color_count(decoded: DecodedPacket) -> int | None:
    """Color slots recoverable from hex. D2 is 0x55 RGB groups; 0F/0E is not unique."""
    fmt, rest = _payload_after_format(decoded)
    if fmt == 0xD2:
        colors, _consumed = _take_d2_colors(rest, None)
        return len(colors) if colors else None
    return None


def _take_d2_colors(rest: list[int], color_count: int | None) -> tuple[list[dict], int]:
    """Walk D2 color slots: `55 r g b` repeats, or `D2 58 r g b` as a second RGB tag."""
    out: list[dict] = []
    i = 0
    limit = int(color_count) if color_count is not None else 8
    while i < len(rest) and len(out) < limit:
        if rest[i] == 0x55 and i + 4 <= len(rest):
            out.append({
                "r": int(rest[i + 1]), "g": int(rest[i + 2]), "b": int(rest[i + 3]),
                "name": "rgb",
            })
            i += 4
            continue
        if rest[i] == 0xD2 and i + 5 <= len(rest) and rest[i + 1] in (0x55, 0x58):
            out.append({
                "r": int(rest[i + 2]), "g": int(rest[i + 3]), "b": int(rest[i + 4]),
                "name": "rgb",
            })
            i += 5
            continue
        if rest[i] == 0x58 and i + 4 <= len(rest) and out:
            out.append({
                "r": int(rest[i + 1]), "g": int(rest[i + 2]), "b": int(rest[i + 3]),
                "name": "rgb",
            })
            i += 4
            continue
        break
    return out, i


def infer_expected_colors(decoded: DecodedPacket, color_count: int | None) -> list[dict]:
    """Palette / D2 RGB slots from hex. 0F/0E needs color_count to know where the tail starts."""
    from .palette import palette_entry

    fmt, rest = _payload_after_format(decoded)
    if fmt == 0xD2:
        colors, _consumed = _take_d2_colors(rest, color_count)
        return colors
    if fmt in (0x0F, 0x0E) and color_count:
        n = max(0, int(color_count))
        return [palette_entry(rest[i] & 0x1F) for i in range(min(n, len(rest)))]
    return []


def tail_bytes_from_decoded(decoded: DecodedPacket, color_count: int | None) -> list[tuple[int, str]]:
    """Post-color-block bytes (minus vib) as T00.. pairs. Used when the sheet has no T columns."""
    fmt, rest = _payload_after_format(decoded)
    n_color = 1 if color_count is None else max(int(color_count), 0)
    if fmt in (0x0F, 0x0E):
        rest = rest[n_color:]
    elif fmt == 0xD2:
        _colors, consumed = _take_d2_colors(rest, color_count)
        rest = rest[consumed:]
    return [(idx, f"{b:02X}") for idx, b in enumerate(rest)]


def trial_from_dict(data: dict[str, Any], *, source_kind: str = SOURCE_BUILDER) -> TrialRow:
    hex_raw = str(data.get("hex_full") or data.get("hex") or "")
    full, env_assumed = ensure_hex_full(hex_raw)
    decoded = decode_hex_structure(full or hex_raw)
    raw_tail = data.get("tail_bytes") or []
    tail: list[tuple[int, str]] = []
    for item in raw_tail:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            tail.append((int(item[0]), str(item[1])))
        elif isinstance(item, dict):
            tail.append((int(item.get("index", 0)), str(item.get("hex", ""))))
    n_colors = data.get("color_count")
    if n_colors is None and decoded:
        n_colors = infer_color_count(decoded)
    if not tail and decoded:
        tail = tail_bytes_from_decoded(decoded, n_colors)
    expected = list(data.get("expected_colors") or [])
    if not expected and decoded:
        expected = infer_expected_colors(decoded, n_colors)
    hint_raw = data.get("zone_layout_hint") or {}
    hint = ZoneLayoutHint(
        five_zones=hint_raw.get("five_zones"),
        sync=hint_raw.get("sync"),
        layout=hint_raw.get("layout"),
        direction=hint_raw.get("direction"),
        n_zones=hint_raw.get("n_zones"),
        cycle_length=hint_raw.get("cycle_length"),
        n_cycles=hint_raw.get("n_cycles"),
    ) if isinstance(hint_raw, dict) else ZoneLayoutHint()
    notes = list(data.get("notes") or [])
    if env_assumed:
        notes.append("envelope_assumed: prepended 8301E100 (payload-only hex)")
    sheet = str(data.get("sheet") or "builder")
    row_index = int(data.get("row_index") or 0)
    return TrialRow(
        sheet=sheet,
        row_id=str(data.get("row_id") or f"{sheet}:{row_index}"),
        row_index=row_index,
        op_code=data.get("op_code"),
        length_byte=data.get("length_byte") if data.get("length_byte") is not None else decoded.length_byte,
        derived_payload_length=data.get("derived_payload_length") if data.get("derived_payload_length") is not None else decoded.derived_payload_length,
        color_count=n_colors,
        color_format_1=data.get("color_format_1"),
        color_format_2=data.get("color_format_2"),
        effect_label=data.get("effect_label"),
        description=data.get("description"),
        hex_full=full if env_assumed else hex_raw,
        hex_key=full,
        location=data.get("location"),
        show=data.get("show"),
        date=data.get("date"),
        tail_start_index=data.get("tail_start_index"),
        tail_bytes=tail,
        vibration_byte=data.get("vibration_byte"),
        notes=notes,
        decoded=decoded,
        zone_layout_hint=hint,
        source_sheet_kind=str(data.get("source_sheet_kind") or source_kind),
        envelope_assumed=env_assumed or bool(data.get("envelope_assumed")),
        expected_colors=expected,
    )


def load_builder_trials(path: str | Path) -> TrialSet:
    """Load one JSON object, a JSON array, or a directory of *.json TrialRow records."""
    import json

    p = Path(path)
    files: list[Path] = []
    if p.is_dir():
        files = sorted(p.glob("*.json"))
    elif p.is_file():
        files = [p]
    else:
        files = sorted(Path().glob(str(path)))
    trials: list[TrialRow] = []
    for f in files:
        data = json.loads(f.read_text(encoding="utf-8"))
        rows = data if isinstance(data, list) else [data]
        for row in rows:
            trials.append(trial_from_dict(row, source_kind=SOURCE_BUILDER))
    return _retag(trials)
